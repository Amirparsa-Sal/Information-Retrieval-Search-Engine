from engine import perform_linguistic_preprocessing, read_dic_from_file, fill_column_numbers
import openpyxl
import pickle
import math
import numpy as np
from numpy.linalg import norm
from gensim.models import Word2Vec
import os
import multiprocessing
import time 

EXCEL_FILE_NAME = 'Data/Merged.xlsx'
wb = openpyxl.load_workbook(EXCEL_FILE_NAME)
sheet = wb.active
MODEL_FILE = 'insa_news.model'

def create_tokens_list(sheet, content_column, delete_stop_words=True):
    doc_token_list = []
    # Looping over the excel file
    for i in range(2, sheet.max_row + 1):
        content = sheet.cell(row=i, column=content_column).value
        # Performing linguistic preprocessing on the content
        token_list = perform_linguistic_preprocessing(content, delete_stop_words=delete_stop_words)
        only_tokens_list = [item[0] for item in token_list]
        doc_token_list.append(only_tokens_list)
    pickle.dump(doc_token_list, open("token_list.obj", "wb"))
    return doc_token_list

def create_token_list_with_count(sheet, content_column, delete_stop_words=True):
    doc_token_list_with_counts = []
    # Looping over the excel file
    for i in range(2, sheet.max_row + 1):
        content = sheet.cell(row=i, column=content_column).value
        # Performing linguistic preprocessing on the content
        token_list = perform_linguistic_preprocessing(content, delete_stop_words=delete_stop_words)
        token_list = [item[0] for item in token_list]
        token_list_with_count = dict(zip(token_list, [token_list.count(item) for item in token_list]))
        doc_token_list_with_counts.append(token_list_with_count)
    pickle.dump(doc_token_list_with_counts, open("token_list_count.obj", "wb"))
    return doc_token_list_with_counts


def create_token_count_dic(docs_token_list_with_count):
    token_count_dic = dict()
    for doc in docs_token_list_with_count:
        for token in doc:
            if token in token_count_dic:
                token_count_dic[token] += doc[token]
            else:
                token_count_dic[token] = doc[token]
    pickle.dump(token_count_dic, open("token_count.obj", "wb"))
    return token_count_dic

def create_tf_idf_list(doc_number, doc_token_list_with_counts, token_count):
    for dic in doc_token_list_with_counts:
        for key in dic:
            dic[key] = (1 + math.log10(dic[key])) * math.log10(doc_number / token_count[key])
    pickle.dump(doc_token_list_with_counts, open("tf_idf_dic.obj", "wb"))
    return doc_token_list_with_counts

def create_docs_matrix(model, docs_tf_id_list):
    matrix = np.array(np.zeros((300, len(docs_tf_id_list))))
    for i in range(len(docs_tf_id_list)):
        vector = np.zeros((300))
        weight_sum = 0
        for key in docs_tf_id_list[i]:
            if key in model.wv:
                vector += model.wv[key] * docs_tf_id_list[i][key]
                weight_sum += docs_tf_id_list[i][key]
        if weight_sum != 0:
            vector /= weight_sum
        for j in range(300):
            matrix[j, i] = vector[j]
    return matrix

def similarity(matrix, doc_1, doc_2):
    score = np.dot(matrix[:, doc_1 - 1].T, matrix[:, doc_2 - 1]) / (norm(matrix[:, doc_1 - 1]) * norm(matrix[:, doc_2 - 1]))
    return (score + 1) / 2

def create_query_vector(model, query, token_count_dic, doc_number):
    # Calculate tf-idf for the query
    query_tf_idf_dic = dict(zip(query, [query.count(item) for item in query]))
    for key in query_tf_idf_dic:
        if key in token_count_dic:
            query_tf_idf_dic[key] = (1 + math.log10(query_tf_idf_dic[key])) * math.log10(doc_number / token_count_dic[key])

    query_vector = np.zeros((300))
    weight_sum = 0
    for key in query:
        if key in model.wv:
            query_vector += model.wv[key] * query_tf_idf_dic[key]
            weight_sum += query_tf_idf_dic[key]
    if weight_sum != 0:
        query_vector /= weight_sum
    return query_vector

def search(model, matrix, query, token_count_dic, doc_number):
    query_vector = create_query_vector(model, query, token_count_dic, doc_number)
    scores = [0 for i in range(matrix.shape[1])]
    for i in range(matrix.shape[1]):
        if norm(matrix[:, i]) != 0:
            scores[i] += np.dot(query_vector.T, matrix[:, i]) / (norm(matrix[:, i]) * norm(query_vector))
    # Sort the scores in descending order
    scores = sorted(enumerate(scores), key = lambda x: x[1], reverse=True)
    return [(str(score[0] + 1), score[1]) for score in scores[:min(10, len(scores))]]

if __name__ == '__main__':
    columns_dic = fill_column_numbers(sheet)
    
    docs_token_list_with_counts = dict()
    if not os.path.exists("token_list_count.obj"):
        print("Creating token list with count")
        docs_token_list_with_counts = create_token_list_with_count(sheet, columns_dic['content'], delete_stop_words=True)
    else:
        docs_token_list_with_counts = pickle.load(open("token_list_count.obj", "rb"))

    token_count_dic = dict()
    if not os.path.exists('token_count.obj'):
        print("Creating token count")
        token_count_dic = create_token_count_dic(docs_token_list_with_counts)
    else:
        token_count_dic = pickle.load(open("token_count.obj", "rb"))
    
    tf_idf_list = []
    if not os.path.exists("tf_idf_dic.obj"):
        print("Creating tf-idf list")
        tf_idf_list = create_tf_idf_list(len(docs_token_list_with_counts), docs_token_list_with_counts, token_count_dic)
    else:
        tf_idf_list = pickle.load(open("tf_idf_dic.obj", "rb"))


    doc_token_list = []
    if not os.path.exists("token_list.obj"):
        print("Creating token list")
        doc_token_list = create_tokens_list(sheet, columns_dic['content'])
    else:
        doc_token_list = pickle.load(open("token_list.obj", "rb"))

    model = None
    if not os.path.exists(MODEL_FILE):
        model = Word2Vec(min_count = 1,
                            window = 5,
                            vector_size = 300,
                            alpha = 0.03,
                            workers = multiprocessing.cpu_count() - 1)
        model.build_vocab(doc_token_list)
        print('Start Training model...')
        start = time.time()
        model.train(doc_token_list, total_examples = model.corpus_count, epochs = 30)
        end = time.time()
        print(f'Completed in {(end - start)} s')
        model.save(MODEL_FILE)
        print(f'Model saved')
    else:
        print("Loading model")
        model = Word2Vec.load(MODEL_FILE)
    
    del doc_token_list

    matrix = create_docs_matrix(model, tf_idf_list)
    
    del tf_idf_list
    del docs_token_list_with_counts

    while True:
        query = input('Enter your query: ')
        query = list(map(lambda token: token[0], perform_linguistic_preprocessing(query)))
        if len(query) != 0:
            start_time = time.time()
            news = search(model, matrix, query, token_count_dic, sheet.max_row - 1)
            end_time = time.time()
            if news is None:
                print('No news found')
            else:
                print('News found:')
                news = list(map(lambda new: (sheet.cell(row=int(new[0])+1, column=columns_dic['url']).value, sheet.cell(row=int(new[0])+1, column=columns_dic['title']).value, new[0], new[1]), news))
                for i,new in enumerate(news):
                    print(new[2] + ': ' + new[1])
                    print(f"Similarity: {new[3]}")
                    print(new[0])
                    print()
                print(f"Time taken: {end_time - start_time}")
        else:
            print('No news found')