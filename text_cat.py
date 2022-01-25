import pickle
from engine import fill_column_numbers, create_index, read_dic_from_file, create_champions_list, create_length_arr, \
    perform_linguistic_preprocessing, search
import openpyxl
import os
import time 
import math

DATA_EXCEL_FILE_NAME = 'Merged.xlsx'
TEST_EXCEL_FILE_NAME = 'data.xlsx'

index = dict() #token -> [freq, {doc_id1: [freq, pos1, pos2, ...], doc_id2: [freq, pos1, pos2, ...], ...}, doc_freq]
EXCEL_FILE_NAME = 'Merged.xlsx'
CHAMPIONS_LIST_SIZE = 20
champions_list = dict()
k = 15

def search_in_topics(index, query, cat, topics, doc_number, length_arr):
    '''A function to perform a ranked retrieval search.'''
    query_items = set([(term,query.count(term)) for term in query])
    # Filtering the documents that do not have the coverage threshold
    all_docs_in_cat = topics[cat]
    print(max(all_docs_in_cat))
    scores = [0 for i in range(len(all_docs_in_cat))]
    print(query_items)
    # Loop over the query items
    for item in query_items:
        # Check if the term is in the index
        if item[0] in index:
            # Calculate tf.idf for each query item in query
            w_tq = (1 + math.log10(item[1])) * math.log10(doc_number / index[item[0]][2])
            # Add tf.id for each query item and its documents
            for i,doc_id in enumerate(all_docs_in_cat):
                if doc_id in index[item[0]][1]:
                    w_dt = (1 + math.log10(index[item[0]][1][doc_id][0])) * math.log10(doc_number / index[item[0]][2])
                    tf_idf = w_tq * w_dt
                    scores[i] += tf_idf
    # devide each score by document length
    for i in range(len(all_docs_in_cat)):
        if length_arr[int(all_docs_in_cat[i]) - 1] != 0:
            scores[i] /= length_arr[int(all_docs_in_cat[i]) - 1]
    # sort the scores
    scores_merged = [(all_docs_in_cat[i], scores[i]) for i in range(len(all_docs_in_cat))]
    scores_merged = sorted(scores_merged, key= lambda x: x[1], reverse=True)
    # return the top 10 results
    return scores_merged[:10]

if __name__ == '__main__':
    wb_data = openpyxl.load_workbook(DATA_EXCEL_FILE_NAME)
    sheet_data = wb_data.active

    data_file_columns_dic = fill_column_numbers(sheet_data)

    if not os.path.exists('index.json'):
        print('Creating index...')
        create_index(sheet_data, index, data_file_columns_dic['content'], delete_stop_words=True)
    else:
        print('Reading index...')
        index = read_dic_from_file('index.json')

    if not os.path.exists(f'champions{CHAMPIONS_LIST_SIZE}.json'):
        print('Creating champions list...')
        champions_list = create_champions_list(index, CHAMPIONS_LIST_SIZE)
    else:
        print('Reading champions list...')
        champions_list = read_dic_from_file(f'champions{CHAMPIONS_LIST_SIZE}.json')

    print('Creating length array...')
    length_arr = create_length_arr(sheet_data, index)

    data_max_row = sheet_data.max_row - 1

    wb_test = openpyxl.load_workbook(TEST_EXCEL_FILE_NAME)
    sheet_test = wb_test.active
    test_file_columns_dic = fill_column_numbers(sheet_test)

    docs_topics = dict()
    for i in range(2, data_max_row + 2):
        topic = sheet_data.cell(row=i, column=data_file_columns_dic['topic']).value
        docs_topics[f'd{i-1}'] = topic

    if not os.path.exists('docs_topics.obj'):
        print("Updating topics...")
        result = dict()
        start_time = time.time()
        for i in range(2, sheet_test.max_row + 1):
            query = sheet_test.cell(row=i, column=test_file_columns_dic['content']).value
            query = list(map(lambda token: token[0], perform_linguistic_preprocessing(query)))
            if len(query) != 0:
                news_list = search(index, query, data_max_row, length_arr, k=k, ranked=True, index_eliminiation_threshold=0.6, doc_coverage_threshold=0.0,champions_list=champions_list, use_champions_list=True)
                topics_count = []
                for news in news_list:
                    topics_count.append(docs_topics[f'd{news}'])
                detected_topic = max(set(topics_count), key=topics_count.count)
                if detected_topic == 'sport':
                    detected_topic = 'sports'
                elif detected_topic == 'political':
                    detected_topic = 'politics'
                docs_topics[f't{i-1}'] = detected_topic
                if detected_topic in result:
                    result[detected_topic] += 1
                else:
                    result[detected_topic] = 1
        end_time = time.time()
        print(result)
        print(end_time - start_time)
        pickle.dump(docs_topics, open('docs_topics.obj', 'wb'))
    else:
        docs_topics = pickle.load(open('docs_topics.obj', 'rb'))

    if not os.path.exists('topics.obj'):
        topics = dict()
        for doc in docs_topics:
            if doc.startswith('t'):
                if docs_topics[doc] in topics:
                    topics[docs_topics[doc]].append(doc[1:])
                else:
                    topics[docs_topics[doc]] = [doc[1:]]
        pickle.dump(topics, open('topics.obj', 'wb'))
    else:
        topics = pickle.load(open('topics.obj', 'rb'))

    del index

    test_index = dict()
    if not os.path.exists('test_index.json'):
        print('Creating test index...')
        create_index(sheet_test, test_index, test_file_columns_dic['content'], delete_stop_words=True, file_name='test_index.json')
    else:
        print('Reading index...')
        test_index = read_dic_from_file('test_index.json')

    print('Creating length array...')
    length_arr = create_length_arr(sheet_test, test_index)

    print(len(length_arr))

    while True:
        query = input('Enter your query: ')
        query = query.split('cat:')
        cat = query[1]
        query = list(map(lambda token: token[0], perform_linguistic_preprocessing(query[0])))
        if len(query) != 0:
            start_time = time.time()
            news = search_in_topics(test_index, query, cat, topics, sheet_test.max_row - 1, length_arr)
            end_time = time.time()
            if news is None:
                print('No news found')
            else:
                print('News found:')
                news = list(map(lambda tuple: (sheet_test.cell(row=int(tuple[0])+1, column=test_file_columns_dic['url']).value,
                                               sheet_test.cell(row=int(tuple[0])+1, column=test_file_columns_dic['title']).value,
                                               tuple[0], tuple[1]), news))
                for i,new in enumerate(news):
                    print(new[2] + ': ' + new[1])
                    print(new[0])
                    print(new[3])
            print("Time taken: ", end_time - start_time)
        else:
            print('No news found')