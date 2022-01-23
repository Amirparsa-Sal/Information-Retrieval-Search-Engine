import multiprocessing
from multiprocessing.spawn import import_main_path
from scipy.fftpack import ss_diff
from word_embedding import create_tf_idf_list, create_docs_matrix, fill_column_numbers, \
    create_token_list_with_count,create_token_count_dic, create_query_vector
from engine import perform_linguistic_preprocessing
import pickle
import numpy as np
from gensim.models import Word2Vec
import openpyxl
import os
import time
from numpy.linalg import norm

EXCEL_FILE_NAME = 'Merged.xlsx'
wb = openpyxl.load_workbook(EXCEL_FILE_NAME)
sheet = wb.active
MODEL_FILE = 'insa_news.model'

def perform_clustering(matrix, k, max_epochs=None):
    distances = np.zeros((k, matrix.shape[1]))
    clusters = np.zeros((1, matrix.shape[1]))
    centroid_numbers = np.random.randint(0, matrix.shape[1], size=(1, k))
    centroids = np.zeros((matrix.shape[0], k))
    for i in range(k):
        centroids[:, i] = matrix[:, centroid_numbers[0, i]]
    change = True
    epoch = 0
    while change and (max_epochs is None or epoch < max_epochs):
        for i in range(k):
            centroid_vector = centroids[:, i].reshape(-1, 1)
            difference = matrix - centroid_vector
            difference = np.square(difference).sum(axis=0)
            distances[i] = difference
        new_clusters = np.argmin(distances, axis=0)
        print(new_clusters)
        new_centroids = np.zeros((matrix.shape[0], k))
        for i in range(k):
            new_centroids[:, i] = np.mean(matrix[:, new_clusters == i], axis=1)
        change = np.count_nonzero(new_clusters - clusters)
        centroids = new_centroids
        clusters = new_clusters
        print(change)
        epoch += 1
    clusters_content = dict()
    for i in range(matrix.shape[1]):
        if clusters[i] in clusters_content:
            clusters_content[clusters[i]].append(i)
        else:
            clusters_content[clusters[i]] = [i]
    return clusters_content, centroids

def search(model, matrix, query, token_count_dic, clusters_content, centroids, doc_number, b = 1):
    query_vector = create_query_vector(model, query, token_count_dic, doc_number).reshape(-1,1)
    nearest_centroids = np.argsort(np.sum(np.square(centroids - query_vector), axis=0))[:b]
    print(f"Searching in clusters {nearest_centroids}")
    docs = []
    for centroid in nearest_centroids:
        for doc_id in clusters_content[centroid]:
            docs.append([doc_id,0])
    for doc in docs:
        if norm(matrix[:, doc[0]]) != 0:
            doc[1] += np.dot(query_vector.T, matrix[:, doc[0]]) / (norm(matrix[:, doc[0]]) * norm(query_vector))
    # Sort the scores in descending order
    docs = sorted(docs, key = lambda x: x[1], reverse=True)
    return [(str(doc[0] + 1), doc[1]) for doc in docs[:min(10, len(docs))]]

if __name__ == '__main__':
    columns_dic = fill_column_numbers(sheet)
    
    docs_token_list_with_counts = dict()
    if not os.path.exists("token_list_count.obj"):
        print("Creating token list with count")
        docs_token_list_with_counts = create_token_list_with_count(sheet, columns_dic['content'], delete_stop_words=True)
    else:
        docs_token_list_with_counts = pickle.load(open("token_list_count.obj", "rb"))

    token_count_dic = dict()
    if not os.path.exists('token_count.obj'):
        print("Creating token count")
        token_count_dic = create_token_count_dic(docs_token_list_with_counts)
    else:
        token_count_dic = pickle.load(open("token_count.obj", "rb"))
    
    tf_idf_list = []
    if not os.path.exists("tf_idf_dic.obj"):
        print("Creating tf-idf list")
        tf_idf_list = create_tf_idf_list(len(docs_token_list_with_counts), docs_token_list_with_counts, token_count_dic)
    else:
        tf_idf_list = pickle.load(open("tf_idf_dic.obj", "rb"))


    docs_token_list = [key for key in token_count_dic]

    model = None
    if not os.path.exists(MODEL_FILE):
        model = Word2Vec(min_count = 1,
                            window = 5,
                            vector_size = 300,
                            alpha = 0.03,
                            workers = multiprocessing.cpu_count() - 1)
        model.build_vocab(docs_token_list)
        print('Start Training model...')
        start = time.time()
        model.train(docs_token_list, total_examples = model.corpus_count, epochs = 30)
        end = time.time()
        print(f'Completed in {(end - start)} s')
        model.save(MODEL_FILE)
        print(f'Model saved')
    else:
        print("Loading model")
        model = Word2Vec.load(MODEL_FILE)
    
    del docs_token_list

    matrix = None
    if not os.path.exists('docs_matrix.obj'):
        print("Creating docs matrix...")
        matrix = create_docs_matrix(model, tf_idf_list)
        np.save('docs_matrix.obj', matrix)
    else:
        print("Loading docs matrix...")
        matrix = np.load('docs_matrix.obj')

    del tf_idf_list
    del docs_token_list_with_counts

    choice = int(input("Do you want to perform clustering? (1 for yes, 0 for no): "))
    clusters_content = None
    centroids = None
    if choice or not (os.path.exists('centroids.obj') and os.path.exists('clusters_content.obj')):
        k = int(input("Enter number of clusters: "))
        clusters_content, centroids = perform_clustering(matrix, k=k)
        pickle.dump(clusters_content, open("clusters_content.obj", "wb"))
        np.save(open("centroids.obj", "wb"), centroids)
    else:
        clusters_content = pickle.load(open("clusters_content.obj", "rb"))
        centroids = np.load(open("centroids.obj","rb"), allow_pickle=True)
    
    while True:
        choice = int(input('What do you want to do?\n1.Search\n2.Get Clusters '))
        if choice == 1:
            query = input("Enter your query: ")
            query = list(map(lambda token: token[0], perform_linguistic_preprocessing(query)))
            if len(query) != 0:
                start_time = time.time()
                news = search(model, matrix, query, token_count_dic, clusters_content, centroids, sheet.max_row - 1, b = 3)
                end_time = time.time()
                if news is None:
                    print('No news found')
                else:
                    print('News found:')
                    news = list(map(lambda new: (sheet.cell(row=int(new[0])+1, column=columns_dic['url']).value, sheet.cell(row=int(new[0])+1, column=columns_dic['title']).value, new[0], new[1]), news))
                    for i,new in enumerate(news):
                        print(new[2] + ': ' + new[1])
                        print(f"Similarity: {new[3]}")
                        print(new[0])
                        print()
            else:
                print('No news found')
            print(f"Time taken: {(end_time - start_time)} s")
            
        elif choice == 2:
            print('Clusters:')
            for cluster_num in clusters_content:
                print(f"Cluster {cluster_num}: {len(clusters_content[cluster_num])} documents")
            cluster_num = -1
            while cluster_num < 0 or cluster_num >= len(clusters_content.keys()):
                cluster_num = int(input("which cluster you want to see? "))
            docs_num = -1
            while docs_num <= 0 or docs_num > len(clusters_content[cluster_num]):
                docs_num = int(input("How many docs you want to see? "))

            docs = clusters_content[cluster_num][:docs_num]
            for i, doc in enumerate(docs):
                print(f"{i}){sheet.cell(row=int(doc)+1, column=columns_dic['title']).value}")
                print(f"{sheet.cell(row=int(doc)+1, column=columns_dic['url']).value}")
                print()