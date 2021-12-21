from __future__ import unicode_literals
from hazm import Normalizer, word_tokenize, stopwords_list, Stemmer
import string
import openpyxl
from parsivar import FindStems
import re
from collections import OrderedDict
import json
import os
import matplotlib.pyplot as plt
import math

stop_words = stopwords_list()
normalizer = Normalizer()
stemmer = FindStems()
index = dict() #token -> [freq, {doc_id1: [freq, pos1, pos2, ...], doc_id2: [freq, pos1, pos2, ...], ...}, doc_freq]
length_arr = []
EXCEL_FILE_NAME = 'data.xlsx'
LINK_REGEX = r"((http|https)\:\/\/)?[a-zA-Z0-9\.\/\?\:@\-_=#]+\.([a-zA-Z]){2,6}([ا-یa-zA-Z0-9\.\&\/\?\:@\-_=# ])*"
punctuations = string.punctuation
punctuations += ''.join(['،','؛','»','«'])
CHAMPIONS_LIST_SIZE=  20
champions_list = dict()

def perform_linguistic_preprocessing(text, delete_stop_words=True, stemming=True):
    '''
    A function to perform linguistic preprocessing on a given text.\n
    Args:\n
        text (str): The text to be processed.\n
        delete_stop_words (bool): A boolean value indicating whether to delete stop words.\n
        stemming (bool): A boolean value indicating whether to perform stemming.\n 
    '''
    #Removing links from the text
    text = re.sub(LINK_REGEX, '', text) 
    #Removing punctuations from the text
    text = text.translate(str.maketrans('', '', punctuations)) 
    #Normalizing the text
    text = normalizer.normalize(text)
    #Tokenizing the text 
    token_list = word_tokenize(text) 
    token_positions = []
    if delete_stop_words:
        for i,token in enumerate(token_list):
            if token not in stop_words:
                token_positions.append([token, i+1])
    else:
        for i,token in enumerate(token_list):
                token_positions.append([token, i+1])
    #Stemming the text
    if stemming:
        return list(map(lambda token: (stemmer.convert_to_stem(token[0]), token[1]), token_positions))

    return token_positions

def create_index(sheet, index, delete_stop_words=True):
    '''
    A function to create an index from the excel file.\n
    Args:\n
        index (dict): The index to be created.\n
        delete_stop_words (bool): A boolean value indicating whether to delete stop words.\n
    '''
    # Looping over the excel file
    for i in range(2, sheet.max_row + 1):
        content = sheet.cell(row=i, column=1).value
        # Performing linguistic preprocessing on the content
        token_list = perform_linguistic_preprocessing(content, delete_stop_words=delete_stop_words)
        # Looping over the tokens
        for token in token_list:
            word = token[0]
            pos = token[1]
            # Checking if the term is already in the index
            if word in index:
                # Increment term frequency
                index[word][0] += 1
                # Checking if the doc_id is already in the postings
                if (i-1) in index[word][1]:
                    # Adding the doc_id to postings and increment term frequency in doc
                    index[word][1][i-1][0] += 1
                    index[word][1][i-1][1].append(pos)
                else:
                    # Adding the new doc_id posting list and increment doc freq
                    index[word][1][i-1] = [1, [pos]]
                    index[word][2] += 1
            else:
                ordered_dict = OrderedDict()
                ordered_dict[i-1] = [1, [pos]]
                index[word] = [1, ordered_dict, 1]

    # Sorting the index Alphabetically
    index = OrderedDict(sorted(index.items(), key=lambda x: x[0], reverse=False))
    # Saving the index to a json file
    json.dump(index, open('index.json', 'w'))

def create_length_arr(sheet, index):
    '''
    A function to calculate length of the document vectors.\n
    Args:\n
        index (dict): The index to be created.\n
    '''
    doc_number = sheet.max_row - 1
    arr = [0 for i in range(doc_number)]
    for term, postings in index.items():
        for doc_id, posting in postings[1].items():
            arr[int(doc_id) - 1] += ((1 + math.log10(posting[0])) * math.log10(doc_number / postings[2])) ** 2
    for i in range(len(arr)):
        arr[i] = math.sqrt(arr[i])
    return arr

def create_champions_list(index):
    champions_list = dict()
    for term, postings in index.items():
        tops = sorted(postings[1].items(), key=lambda x: x[1][0], reverse=True)
        top_k = tops[:min(CHAMPIONS_LIST_SIZE,len(tops))]
        champions_list[term] = [item[0] for item in top_k]
    json.dump(champions_list, open(f'champions{CHAMPIONS_LIST_SIZE}.json', 'w'))
    return champions_list
    
def read_dic_from_file(file_name):
    '''A function to read dictionary from a json file.'''
    value = json.loads(open(file_name, 'r').read())
    return OrderedDict(value.items())

def single_word_query(word):
    '''A function to perform a single word boolean query.'''
    # Checking if the word is in the index
    if word in index:
        # Getting the posting list
        news = []
        for doc_id in index[word][1]:
            news.append(doc_id)
        return news
    else:
        return []

def has_positions(postings1, doc_id1, postings2, doc_id2):
    '''A function to check if two posting list have 2 consecutive positions.'''
    pos_pointer1 = 0
    pos_pointer2 = 0
    pos_list1 = postings1[1][doc_id1][1]
    pos_list2 = postings2[1][doc_id1][1]
    while pos_pointer1 < len(pos_list1) and pos_pointer2 < len(pos_list2):
        pos1 = pos_list1[pos_pointer1]
        pos2 = pos_list2[pos_pointer2]
        if pos1 == pos2 - 1:
            return True
        elif pos1 < pos2 - 1:
            pos_pointer1 += 1
        else:
            pos_pointer2 += 1
    return False

def intersect(list1, list2):
    '''A function to perform an intersection between two posting lists.'''
    i = 0
    j = 0
    result = []
    while i < len(list1) and j < len(list2):
        if list1[i] == list2[j]:
            result.append(list1[i])
            i += 1
            j += 1
        elif list1[i] < list2[j]:
            i += 1
        else:
            j += 1
    return result

def combine_results(results1, results2):
    '''A function to combine two posting lists.'''
    for r2 in results2:
        if r2 not in results1:
            results1.append(r2)

def double_word_query(word1, word2):
    '''A function to perform a double word boolean query.'''
    postings1 = index.get(word1, None)
    postings2 = index.get(word2, None)
    # Checking if the words are in the index
    if postings1 is not None and postings2 is not None:
        news = []
        doc_pointer1 = 0
        doc_pointer2 = 0
        doc_id_list1 = list(postings1[1].keys())
        doc_id_list2 = list(postings2[1].keys())
        # Looping over the posting lists
        while doc_pointer1 < len(doc_id_list1) and doc_pointer2 < len(doc_id_list2):
            doc_id1 = doc_id_list1[doc_pointer1]
            doc_id2 = doc_id_list2[doc_pointer2]
            # Checking if the doc_ids are equal
            if doc_id1 == doc_id2:
                # Checking if we have 2 consecutive positions
                if has_positions(postings1, doc_id1, postings2, doc_id2):
                    news.append(doc_id1)
                doc_pointer2 += 1
                doc_pointer1 += 1
            elif int(doc_id1) < int(doc_id2):
                doc_pointer1 += 1
            else:
                doc_pointer2 += 1
        return news
    return []

def multiple_word_query(query):
    '''A function to perform a multiple word boolean query.'''
    # Use single_word_query function if the query is a single word
    if len(query) == 1:
        return single_word_query(query[0])
    else:
        news = []
        # Loop over the length of the different permutations of the query
        for i in range(len(query), 1, -1):
            # Loop over the permutations of the query with the length i
            for j in range(0, len(query) - i + 1): 
                # Get a permutation of the query with the length i
                search_list = query[j:j+i]
                result = []
                # Intersect each double word query within the permutation
                for k in range(len(search_list) - 1): 
                    if result is None:
                        result = double_word_query(search_list[k], search_list[k+1])
                    else:
                        result = intersect(result, double_word_query(search_list[k], search_list[k+1]))
                    if result is None:
                        break
                # Combine the results of the double word queries
                if len(result) > 0:
                    combine_results(news, result)
        # Adding single word results
        for i in range(len(query)):
            combine_results(news, single_word_query(query[i]))

        return news

def doc_has_coverage(doc_id, query_terms, coverage_threshold=0.6):
    total_terms = len(query_terms)
    n = 0
    for term in query_terms:
        if term in index:
            if doc_id in index[term][1]:
                n += 1
    if n / total_terms >= coverage_threshold:
        return True
    return False
    
def ranked_retreival_search(doc_number, query, index_elimination_threshold=0.0, doc_coverage_threshold=0.6, use_champions_list=False):
    '''A function to perform a ranked retrieval search.'''
    query_items = set([(term,query.count(term)) for term in query])
    all_docs_containing_terms = set()
    if not use_champions_list:
        # Adding all documents containing at least one of the query terms (index elimination)
        for item in query_items:
            if item[0] in index and math.log10(doc_number / index[item[0]][2]) >= index_elimination_threshold:
                postings = index[item[0]][1]
                for doc_id in postings:
                    all_docs_containing_terms.add(doc_id)
    else:
        # Adding champions list docs
        for item in query_items:
           if item[0] in index and math.log10(doc_number / index[item[0]][2]) >= index_elimination_threshold:
               all_docs_containing_terms.update(champions_list[item[0]]) 

    # Filtering the documents that do not have the coverage threshold
    all_docs_containing_terms = list(filter(lambda x: doc_has_coverage(x, [item[0] for item in query_items], doc_coverage_threshold), all_docs_containing_terms))

    scores = [0 for i in range(doc_number)]
    # Loop over the query items
    for item in query_items:
        # Check if the term is in the index
        if item[0] in index:
            # Calculate tf.idf for each query item in query
            w_tq = (1 + math.log10(item[1])) * math.log10(doc_number / index[item[0]][2])
            # Add tf.id for each query item and its documents
            for doc_id in all_docs_containing_terms:
                if doc_id in index[item[0]][1]:
                    w_dt = (1 + math.log10(index[item[0]][1][doc_id][0])) * math.log10(doc_number / index[item[0]][2])
                    tf_idf = w_tq * w_dt
                    scores[int(doc_id) - 1] += tf_idf
    # devide each score by document length
    for i in range(doc_number):
        if length_arr[i] != 0:
            scores[i] /= length_arr[i]
    # sort the scores
    scores_sorted = sorted(enumerate(scores), key=lambda x: x[1], reverse=True)
    # return the top 10 results
    return [str(x[0] + 1) for x in scores_sorted[:10]]

def search(query, sheet, ranked=True, index_eliminiation_threshold=0.0, doc_coverage_threshold=0.6, use_champions_list=False):
    if ranked:
        return ranked_retreival_search(sheet.max_row - 1, query, index_elimination_threshold=index_eliminiation_threshold, doc_coverage_threshold=doc_coverage_threshold, use_champions_list=use_champions_list)
    return multiple_word_query(query)

def plot_zipf_law(index):
    '''A function to plot the Zipf law.'''
    index = OrderedDict(sorted(index.items(), key=lambda x: x[1][0], reverse=True))
    word_list = list(index.keys())
    count_multiply_rank = []
    count = []
    for i in range(len(word_list)):
        count_multiply_rank.append(math.log10(index[word_list[i]][0]))
        count.append(index[word_list[i]][0])
    ranks = list(map(lambda x: math.log10(x), range(1, len(word_list)+1)))
    plt.plot(ranks, count_multiply_rank)
    plt.show()

def count_tokens_and_text_length(sheet, n, stemming=True):
    '''A function to count the number of tokens and the length of the text.'''
    tokens = dict()
    length = 0
    for i in range(2, 2 + n):
        content = sheet.cell(row=i, column=1).value
        length += len(word_tokenize(content))
        token_list = perform_linguistic_preprocessing(content, stemming=stemming)
        for token in token_list:
            tokens[token[0]] = True
    return len(tokens), length

if __name__ == '__main__':

    wb = openpyxl.load_workbook(EXCEL_FILE_NAME)
    sheet = wb.active

    if not os.path.exists('index.json'):
        print('Creating index...')
        create_index(sheet, index, delete_stop_words=True)
    else:
        print('Reading index...')
        index = read_dic_from_file('index.json')

    if not os.path.exists(f'champions{CHAMPIONS_LIST_SIZE}.json'):
        print('Creating champions list...')
        champions_list = create_champions_list(index)
    else:
        print('Reading champions list...')
        champions_list = read_dic_from_file(f'champions{CHAMPIONS_LIST_SIZE}.json')

    print('Creating length array...')
    length_arr = create_length_arr(sheet, index)

    while True:
        query = input('Enter your query: ')
        query = list(map(lambda token: token[0], perform_linguistic_preprocessing(query)))
        if len(query) != 0:
            news = search(query, sheet, ranked=True, index_eliminiation_threshold=0.0, doc_coverage_threshold=0.5, use_champions_list=True)
            if news is None:
                print('No news found')
            else:
                print('News found:')
                news = list(map(lambda doc_id: (sheet.cell(row=int(doc_id)+1, column=3).value,doc_id), news))
                for i,new in enumerate(news):
                    print(new[1] + ': ' + new[0])
        else:
            print('No news found')
        
        